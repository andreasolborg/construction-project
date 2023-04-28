import os
from openpyxl import load_workbook, Workbook
from Task import *


class Project:
    def __init__(self, r):
        self.tasks = None
        self.duration = 0
        self.shortest_duration = 0
        self.expected_duration = 0
        self.longest_duration = 0
        self.classification = None
        self.r = r

    def set_shortest_duration(self):
        self.shortest_duration = self.find_early_dates(0)

    def set_expected_duration(self):
        self.expected_duration = self.find_early_dates(1)

    def set_longest_duration(self):
        self.longest_duration = self.find_early_dates(2)
    
    def get_tasks(self):
        return self.tasks
    
    def get_task_by_code(self, code):
        for task in self.tasks:
            if task.code == code:
                return task
        return None
    
    def add_task(self, task):
        self.tasks.append(task)
    
    def __repr__(self):
        return f"{self.classification}"
    
    # Find the critical tasks. A task is critical if its early and late dates are equal.
    def set_is_critical_for_all_tasks(self):
        for task in self.tasks:
            task.set_is_critical()


    # Must be tested
    def is_gate_valid(self, list_of_predecessors):
        # Check if the predecessors are on the same level, i.e. they share the same successor
        successors = [self.get_task_by_code(predecessor).successors for predecessor in list_of_predecessors]
        for successor in successors:
            if successor != successors[0]: # If the successor is not the same as the first successor
                return False
        return True
    
        
        

    def add_gate(self, code, description, list_of_predecessors):
        predecessors = [self.get_task_by_code(predecessor) for predecessor in list_of_predecessors]
        successors = predecessors[0].successors
        
        if self.is_gate_valid(list_of_predecessors):  
            for predecessor in predecessors:
                # Clear the successors of the predecessors
                predecessor.successors = []

            # Create the new gate
            gate = Task("Gate", code, description, [0, 0, 0], predecessors, self.r)

            # Insert the new gate in the list of tasks in the correct position
            index = self.tasks.index(max(predecessors, key=lambda x: self.tasks.index(x)))
            print("Index: {}".format(index))
            self.tasks.insert(index + 1, gate)

            for successor in successors:
                # Clear the predecessors of the successors
                successor.predecessors = []
                # Add the new gate as predecessor to the successors
                successor.add_predecessor(gate)
        else:
            print("Gate is not valid, the predecessors are not on the same level.")

    def get_task_index(self, task):
        return self.tasks.index(task)
        
    def write_task(self, dot_file):
        for task in self.tasks:
            dot_file.write('{} [label="{}"];\n'.format(str(id(task)), str(task.code)))
            for predecessor in task.predecessors:
                dot_file.write('{} -> {};\n'.format(str(id(predecessor)), str(id(task))))

    def draw_pert_diagram(self, filename):
        print("Saving graph to file {}".format(filename)) # print to console
        if os.path.exists("{}.dot".format(filename)): # if file already exists, delete it
            os.remove("{}.dot".format(filename))
        if os.path.exists("{}.png".format(filename)): # if file already exists, delete it
            os.remove("{}.png".format(filename))
        with open("{}.dot".format(filename), "w") as dot_file: 
            dot_file.write("digraph G {\n")
            dot_file.write('rankdir=LR;\ncenter=true;\n')
            self.write_task(dot_file) # print the first task
            dot_file.write("}\n")
        os.system("dot -Tpng -Gdpi=500 {}.dot -o {}.png".format(filename, filename)) # create png from dot file

    # Read the tasks from an excel file
    def import_project_from_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        tasks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            type = row[0]
            code = row[1]
            description = row[2]

            durations = eval(str(row[3]))
            durations = (0.0, 0.0, 0.0) if durations is None else durations

            predecessors = [] if row[4] is None else row[4].split(", ")
            if type == "Task" or type == "Gate":
                tasks.append(Task(type, code, description, durations, predecessors, self.r)) # takes in r from the project as a default value to be used in the task
        
        # Convert the predecessors from a list of codes to a list of tasks
        for task in tasks:
            new_predecessors = []
            for comparing_task in tasks:
                if comparing_task.code in task.predecessors:
                    new_predecessors.append(comparing_task)
                    comparing_task.add_successor(task)
            task.predecessors = new_predecessors
        self.tasks = tasks
        wb.close()

    def export_detailed_project_to_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Type", "Code", "Description", "Min. duration", "Most likely duration", "Max. duration", "Predecessors", "Successors", "Early start date", "Early completion date", "Late start date", "Late completion date", "Is critical"])
        for task in self.tasks:
            ws.append([task.type, task.code, task.description, task.list_of_durations[0], task.list_of_durations[1], task.list_of_durations[2], ", ".join([t.code for t in task.predecessors]), ", ".join([t.code for t in task.successors]), task.early_start_date, task.early_completion_date, task.late_start_date, task.late_completion_date, task.is_critical])
        wb.save(filename)
        wb.close()

    def print_project(self):
        print(f"Tasks: {self.tasks}")
        print("Tasks:")
        for task in self.tasks:
            print(f"Task: {task}")
            print(f"Description: {task.description}")
            print(f"Predecessors: {task.predecessors}")
            print(f"Successors: {task.successors}")
            print(f"Durations: {task.duration}")
            print(f"Early start date: {task.early_start_date}")
            print(f"Early completion date: {task.early_completion_date}")
            print(f"Late start date: {task.late_start_date}")
            print(f"Late completion date: {task.late_completion_date}")
            print(f"Is critical: {task.is_critical}")
            print()

    def find_early_dates(self, duration_index=None): 
        '''
        Early dates are thus calculated by propagating values from the source nodes to the
        sink nodes. The early start date of a task is the maximum of the early completion dates of
        its predecessors plus the task duration. The early completion date of a task is the sum of
        its early start date and its duration.

        1. Start with a list of all tasks
        2. While it remains a task in the list,
            a. Find a task with no predecessors
            b. Remove it from the list
            c. Calculate its early start date and early completion date
        The minimum of the porject is the maximum of the early completion dates of the tasks
        Early dates is thus calculated by propagating values from the source nodes to the sink nodes
        '''

        self.duration = 0
        tasks = self.tasks.copy()
        while len(tasks) > 0:
            for task in tasks:
                if task.has_predecessor_in_list(tasks):
                    continue
                if len(task.predecessors) == 0:
                    task.early_start_date = 0
                    if duration_index is not None:
                        task.early_completion_date = task.early_start_date + task.list_of_durations[duration_index]
                    else:
                        task.early_completion_date = task.early_start_date + task.duration
                else:
                    task.early_start_date = max([predecessor.early_completion_date for predecessor in task.predecessors])
                    if duration_index is not None:
                        task.early_completion_date = task.early_start_date + task.list_of_durations[duration_index]
                    else:
                        task.early_completion_date = task.early_start_date + task.duration
                tasks.remove(task)
        self.duration = max([task.early_completion_date for task in self.tasks])
        return self.duration
        
    def find_late_dates(self, duration_index=None):
        tasks = self.tasks.copy()
        while len(tasks) > 0:
            for task in tasks:
                if task.has_successor_in_list(tasks):
                    continue
                if len(task.successors) == 0:
                    task.late_completion_date = task.early_completion_date
                    task.late_start_date = task.early_start_date
                else:
                    task.late_completion_date = min([successor.late_start_date for successor in task.successors])
                    if duration_index is not None:
                        task.late_start_date = task.late_completion_date - task.list_of_durations[duration_index]
                    else:
                        task.late_start_date = task.late_completion_date - task.duration
                tasks.remove(task)

    def classify_project(self):
        
        '''
        Classify the project as either a success, acceptable, or failure.
        Success: The projects actual duration does not exceed the expected duration by more than 5% (with a risk factor of 1.0)
        Acceptable: The actual duration of the project stands between 105% and 115% of the expected duration (with a risk factor of 1.0)
        Failure: The actual duration of the project exceeds its expected duration by more than 15% (with a risk factor of 1.0)
        '''
        
        if self.duration <= self.expected_duration * 1.05:
            self.classification = "Success"
        elif self.duration <= self.expected_duration * 1.15:
            self.classification = "Acceptable"
        else:
            self.classification = "Failure"
        return self.classification
    

def main():
    project = Project(1.0)
    project.import_project_from_excel("Villa.xlsx")
    project.add_gate("Test_Gate", "Test gate", ["H.2", "H.3"])
    # project.find_early_dates()
    # project.find_late_dates()
    # project.set_expected_duration()
    # project.set_shortest_duration()
    # project.set_longest_duration()

    project.print_project()

    # project.draw_pert_diagram("Villah2h3")

    # Add a gate at the end of the project

    # project.set_is_critical_for_all_tasks()
    # project.classify_project()
    # project.export_detailed_project_to_excel("Villa_output_TD_with_gate.xlsx")

    gate = project.get_task_by_code("Test_Gate")
    idx = project.get_task_index(gate)
    print(idx)
if __name__ == "__main__":
    main()
