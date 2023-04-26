'''
Table 1 describes the tasks of the construction of a warehouse.
Types   | Codes | Description                   | Duration | Predecessors
--------|-------|-------------------------------|----------|--------------
Gate    | Start |                               |          | 
Task    | A     | Agreement of the owner        | 3, 4, 6 | Start
Task    | B     | Preparation of the ground     | 1, 2, 4 | Start
Task    | C     | Order of materials            | 1, 1, 3 | A
Task    | D     | Excavation                    | 1, 1, 3 | A, B
Task    | E     | Order of doors and windows    | 1, 2, 4 | A
Task    | F     | Delivery of materials         | 1, 2, 4 | C
Task    | G     | Concrete foundations          | 1, 2, 4 | D, F
Task    | H     | Delivery of doors and windows | 8, 10, 12 | E
Task    | J     | Frame, roofs, walls           | 3, 4, 6 | G
Task    | K     | Installation of doors and windows | 1, 1, 3 | H, J
Gate    | End   |                               |          | J, H



Table 2: Early and late dates for the warehouse project

Task | Start | A | B | C | D | E | F | G | H | J | K | En
-----|-------|---|---|---|---|---|---|---|---|---|---|-----
Duration | 0 | 4 | 2 | 1 | 2 | 2 | 2 | 2 | 10 | 4 | 1 | 0
Early start date | 0 | 0 | 0 | 4 | 4 | 4 | 5 | 7 | 6 | 9 | 16 | 17
Early completion date | 0 | 4 | 2 | 5 | 5 | 6 | 7 | 9 | 16 | 13 | 17 | 17
Late start date | 0 | 0 | 7 | 7 | 9 | 4 | 8 | 10 | 6 | 12 | 16 | 17
Late completion date | 0 | 4 | 9 | 8 | 10 | 6 | 10 | 12 | 16 | 16 | 17 | 17


'''
import random
from openpyxl import load_workbook
import statistics
import pandas as pd
from sklearn import linear_model, metrics, svm, tree
from sklearn.discriminant_analysis import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split


class Task:
    def __init__(self, type, code, description, durations, predecessors, r=1):  # r is the risk factor, default is 1 (no risk). May be more dynamic in the future
        self.code = code
        self.description = description
        self.type = type

        self.durations = durations
        durations_updated = self.update_durations_with_risk_factor(r, durations)  # Update the durations with the risk factor, First in task 2.2 in the assignment
        self.duration = random.triangular(*durations_updated) # Randomly generate a duration from the given range of durations (min, mode, max)

        # print()
        # print(self.duration)
        # print(self.durations[1])
        # print()

        self.predecessors = predecessors
        # self.duration = durations[1] # Use the mode as the duration
        self.successors = []
        self.early_start_date = 0
        self.early_completion_date = 0
        self.late_start_date = 0
        self.late_completion_date = 0
        self.is_critical = False

    def update_durations_with_risk_factor(self, r, durations):
        a = durations[0] # a = min, b = max, e = mode
        e = durations[1]
        b = durations[2]
        e_new = e * r
        if e_new < a:
            e_new = a
        elif e_new > b:
            # set e_new as b
            e_new = b
        durations = (a, e_new, b)
        return durations

    def add_predecessor(self, predecessor):
        self.predecessors.append(predecessor)
        predecessor.successors.append(self)

    def add_successor(self, successor):
        self.successors.append(successor)
        successor.predecessors.append(self)    
        
    def clear_predecessors(self):
        self.predecessors = []

    def clear_successors(self):
        self.successors = []

    def has_predecessor_in_list(self, tasks):
        for task in tasks:
            if task in self.predecessors:
                return True
        return False
    
    def has_successor_in_list(self, tasks):
        for task in tasks:
            if task in self.successors:
                return True
        return False


    def __str__(self):
        return f"{self.code}"
    
    def __repr__(self):
        return f"{self.code}"


class Project:
    def __init__(self, tasks, r):
        self.tasks = tasks
        self.duration = 0
        self.shortest_duration = 0
        self.expected_duration = 0
        self.longest_duration = 0
        self.classification = None
        self.r = r

    def get_task_by_code(self, code):
        for task in self.tasks:
            if task.code == code:
                return task
        return None

    def set_shortest_duration(self):
        self.shortest_duration = 0
        self.shortest_duration = self.find_early_dates(0)

    def set_expected_duration(self):
        self.expected_duration = 0
        self.expected_duration = self.find_early_dates(1)

    def set_longest_duration(self):
        self.longest_duration = 0
        self.longest_duration = self.find_early_dates(2)
    
    def set_random_duration(self):  # May be removed later
        self.random_duration = 0
        self.random_duration = self.find_early_dates()
    
    def get_tasks(self):
        return self.tasks
    
    def __repr__(self):
        return f"{self.classification}"

    
    # Read the tasks from an excel file
    def import_project_from_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        tasks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            type = row[0]
            code = row[1]
            description = row[2]

            durations = eval(str(row[3]))
            durations = (0.0, 0.0, 0.0) if durations is None else durations

            predecessors = [] if row[4] is None else row[4].split(", ")
            if type == "Task" or type == "Gate":
                tasks.append(Task(type, code, description, durations, predecessors, self.r)) # takes in r from the project as a default value to be used in the task
        # Convert the predecessors from a list of codes to a list of tasks
        for task in tasks:
            new_predecessors = []
            for comparing_task in tasks:
                if comparing_task.code in task.predecessors:
                    new_predecessors.append(comparing_task)
                    comparing_task.add_successor(task)
            task.predecessors = new_predecessors
        self.tasks = tasks
        wb.close()

    def print_project(self):
        '''
        Design a printer to print projects, including the list of successors of each task, their
        early and late dates, and their criticality. Use this printer to test both the data
        structures you designed in Task 1. and your loader.
        '''    
        print(f"Tasks: {self.tasks}")
        print("Tasks:")
        for task in self.tasks:
            print(f"Task: {task}")
            print(f"Description: {task.description}")
            print(f"Predecessors: {task.predecessors}")
            print(f"Successors: {task.successors}")
            print(f"Duration: {task.duration}")
            print(f"Early start date: {task.early_start_date}")
            print(f"Early completion date: {task.early_completion_date}")
            print(f"Late start date: {task.late_start_date}")
            print(f"Late completion date: {task.late_completion_date}")
            print(f"Is critical: {task.is_critical}")
            print()

    def find_early_dates(self, duration_index=None):
        '''
        Early dates are thus calculated by propagating values from the source nodes to the
        sink nodes. The early start date of a task is the maximum of the early completion dates of
        its predecessors plus the task duration. The early completion date of a task is the sum of
        its early start date and its duration.

        1. Start with a list of all tasks
        2. While it remains a task in the list,
            a. Find a task with no predecessors
            b. Remove it from the list
            c. Calculate its early start date and early completion date
        The minimum of the porject is the maximum of the early completion dates of the tasks
        Early dates is thus calculated by propagating values from the source nodes to the sink nodes
        '''
        self.duration = 0
        tasks = self.tasks.copy()
        while len(tasks) > 0:
            for task in tasks:
                # if duration_index is not None:
                #     task.duration = task.durations[duration_index]  # 0: shortest, 1: expected, 2: longest, hvilken foretrekker du?
                if task.has_predecessor_in_list(tasks):
                    continue
                if len(task.predecessors) == 0:
                    task.early_start_date = 0
                    task.early_completion_date = task.early_start_date + task.duration
                else:
                    task.early_start_date = max([predecessor.early_completion_date for predecessor in task.predecessors])
                    if duration_index is not None:
                        task.early_completion_date = task.early_start_date + task.durations[duration_index]
                    else:
                        task.early_completion_date = task.early_start_date + task.duration
                tasks.remove(task)
        self.duration = max([task.early_completion_date for task in self.tasks])
        return self.duration
        
    def find_late_dates(self, duration_index=None):
        tasks = self.tasks.copy()
        while len(tasks) > 0:
            for task in tasks:
                # if duration_index is not None:
                #     task.duration = task.durations[duration_index]
                if task.has_successor_in_list(tasks):
                    continue
                if len(task.successors) == 0:
                    task.late_completion_date = task.early_completion_date
                    task.late_start_date = task.early_start_date
                else:
                    task.late_completion_date = min([successor.late_start_date for successor in task.successors])
                    if duration_index is not None:
                        task.late_start_date = task.late_completion_date - task.durations[duration_index]
                    else:
                        task.late_start_date = task.late_completion_date - task.duration
                tasks.remove(task)
        #         print("Removed task: ", task)
        # print("Late dates found.")

    # Find the critical tasks. A task is critical if its early and late dates are equal.
    # Rename to determine_critical_tasks?
    def find_critical_tasks(self):
        for task in self.tasks:
            if task.early_start_date == task.late_start_date and task.early_completion_date == task.late_completion_date:
                task.is_critical = True
            else:
                task.is_critical = False
            
    
    ######################
    # Machine Learning
    ######################

    def classify_project(self):
        '''
        Classify the project as either a success, acceptable, or failure.
        Success: The projects actual duration does not exceed the expected duration by more than 5% (with a risk factor of 1.0)
        Acceptable: The actual duration of the project stands between 105% and 115% of the expected duration (with a risk factor of 1.0)
        Failure: The actual duration of the project exceeds its expected duration by more than 15% (with a risk factor of 1.0)
        '''
        # Calculate the actual duration of the project
        actual_duration = self.duration
        # Calculate the expected duration of the project
        expected_duration = self.expected_duration
        
        if actual_duration <= expected_duration * 1.05:
            self.classification = "Success"
        elif actual_duration <= expected_duration * 1.15:
            self.classification = "Acceptable"
        else:
            self.classification = "Failure"
        return self.classification
    

    #WIP. I consider we do not add a gate to the project at all, but just do machine learning on the dataset.
    def add_gate(self, code, description, predecessors):
        predecessors = [self.tasks[predecessor] for predecessor in predecessors]
        successors = predecessors[0].successors

        # Clear the successors of the predecessors
        for predecessor in predecessors:
            predecessor.successors = []

        # Create the gate
        gate = Task(type="Gate", code=code, description=description, durations=[0, 0, 0], predecessors=predecessors)
        self.tasks.append(gate)

        for successor in successors:
            self.predecessors = []
            gate.add_successor(successor)
    
    


    
'''
Make at random a sample 1000 of values of durations for each value of the risk factor.
Calculate the duration of the project for each of the 1000 samples.
'''
def make_samples(n):
    risk_factors = [0.8, 1.0, 1.2, 1.4]
    samples_with_risk_factor = {}
    for risk_factor in risk_factors:
        sample_with_risk_factor = []
        for i in range(n):
            project = Project([], risk_factor)
            project.import_project_from_excel("Villa.xlsx")
            project.set_expected_duration()
            project.find_early_dates()
            project.classify_project()
            sample_with_risk_factor.append(project)

        samples_with_risk_factor[risk_factor] = sample_with_risk_factor
    return samples_with_risk_factor

def write_to_csv(samples_with_risk_factor):
    '''
    Takes in a dictionary with risk factors as keys and a list of samples as values. Randomly choose a sample from each risk factor and write it to a csv file.
    '''
    amount_of_samples = len(samples_with_risk_factor[0.8]) # Randomly choose a sample from each risk factor
    for i in range(amount_of_samples):
        random_risk_factor = random.choice(list(samples_with_risk_factor.keys()))
        sample = samples_with_risk_factor[random_risk_factor][i]
        samples_to_save = []
        tasks = []
        task_duration = 0
        for task in sample.tasks[1:-1]: # Exclude the first and last task
            task_duration += task.duration
            tasks.append(task_duration)
        tasks.append(sample.classification)
        samples_to_save.append(tasks)
        #Overwrite the file if it already exists
        pd.DataFrame(samples_to_save).to_csv("samples.csv", mode='a', header=False, index=False)


def machine_learning():
    '''
    Perform machine learning on the csv file. Use the first 80% of the samples to train the model and the last 20% to test the model.
    Use the following algorithms: Logistic Regression, Random Forest, Support Vector Machine
    '''
    # Read the csv file
    df = pd.read_csv("samples.csv", header=None)
    # Split the data into features and labels
    X = df.iloc[:, :-1].values 
    y = df.iloc[:, -1].values
    # Split the data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)
    # Scale the data. This is necessary for SVM, because it uses the euclidean distance.
    scaler = StandardScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    # Train the model
    models = []
    models.append(("LR", linear_model.LogisticRegression()))
    models.append(("RF", RandomForestClassifier()))
    models.append(("SVM", svm.SVC()))
    for name, model in models:
        model.fit(X_train, y_train)
        # Evaluate the model
        y_pred = model.predict(X_test)
        print(name, ":", metrics.classification_report(y_test, y_pred))


    



def perform_statistics(samples_with_risk_factor):
    '''
    Perform statistics on these durations (minimum, maximum,
    mean, standard-deviation, deciles), as well as on the numbers of successful, acceptable
    and failed projects.
    '''
    for risk_factor in samples_with_risk_factor:
        samples = samples_with_risk_factor[risk_factor]
        durations = [sample.duration for sample in samples]
        classifications = [sample.classification for sample in samples]
        print("Risk factor: ", risk_factor)
        # print("Durations: ", durations)
        # print("Classifications: ", classifications)
        print("Minimum duration: ", min(durations))
        print("Maximum duration: ", max(durations))
        print("Mean duration: ", sum(durations)/len(durations))
        print("Standard deviation: ", statistics.stdev(durations))
        print("Deciles: ", statistics.quantiles(durations, n=10))
        print("Number of successes: ", classifications.count("Success"))
        print("Number of acceptables: ", classifications.count("Acceptable"))
        print("Number of failures: ", classifications.count("Failure"))
        print("")

        

    

def main():
    project = Project([], 1.0)
    project.import_project_from_excel("Villa.xlsx")
    project.find_early_dates(1)
    project.find_late_dates(1)
    project.find_critical_tasks()
    # project.print_project()
    # project.print_project()
    print(project.duration, "project duration")

    # project.set_shortest_duration()
    # project.set_expected_duration()
    # project.set_longest_duration()


    # print("Shortest duration: ", project.shortest_duration)
    # print("Expected duration: ", project.expected_duration)
    # print("Longest duration: ", project.longest_duration)

    # samples = make_samples(1000)
    # perform_statistics(samples)

    # write_to_csv(samples)

    machine_learning()



if __name__ == "__main__":
    main()
